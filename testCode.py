from tkinter import filedialog
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from tkinter import *
import csv
from InvoiceGen import InvoiceGen

fileJobCodes = 'job_codes.xlsx'

# try:
#   newInstance = InvoiceGen(['test.xlsx'], fileJobCodes)
# except:
#   print('an error has occurred')
# newInstance.getFileNames()
newInstance = InvoiceGen(['Indeed_itemized_526.93.csv', "Indeed_itemized_526.93.csv"], fileJobCodes)
newInstance.createJobCodeMap()


root = Tk()
root.title('Indeed Extractor')
root.geometry("500x500")

fileName = 'Indeed_itemized_526.93.csv'


jobTitleMap = {}

outputFileName = 'output.xlsx'

outputMap = {}

############ test code ###########
# wb = openpyxl.load_workbook(filename= '')

# f = open('Indeed_itemized_526.93.csv')

# csv_reader_object = csv.reader(f)

# for line in csv_reader_object:
#     print(line)
################ end test code ################

###need to allow multiple csvs to be imported #####
### convert csv file into xlsx file ########
# wb = openpyxl.Workbook()
# ws = wb.active

# with open(fileName) as f:
#     reader = csv.reader(f, delimiter=',')
#     for row in reader:
#         ws.append(row)

# wb.save('file.xlsx')
###### end converting csv file######


###### creates dictionary of job titles as keys and company name and code as value in a dictionary ######
jobCodeWb = openpyxl.load_workbook(filename = fileJobCodes)
jobCodeSheet = jobCodeWb['Sheet1']

# for row_cells in jobCodeSheet.iter_rows(min_row=2):
    # for cell in row_cells:
        # print('%s: cell.value=%s' % (cell, cell.value) )
        # print(cell.value)

for cell in jobCodeSheet['B']:
  # print(cell.coordinate)
  if cell.row == 1 or cell.value.strip() == '':
    continue
  # print(jobTitleMap.get(cell.value, 'None'))
  if(jobTitleMap.get(cell.value, 'None') == 'None'):
    jobTitleMap[cell.value] = {
      'company': jobCodeSheet['A' + str(cell.row)].value,
      'code': jobCodeSheet['C'+ str(cell.row)].value
    }
  # print(cell.value)
  # print(cell.row)
  # print(cell.column)
# print(sheet_ranges['D18'].value)
# print(jobTitleMap)

########## read csv converted file and create new excel wb #################
readWb = openpyxl.load_workbook(filename= 'file.xlsx')
readWbSheet = readWb['Sheet']

invoiceNumber = readWbSheet['A1'].value.split('#')[1]
# print(readWbSheet['A1'].value)
# print(invoiceNumber)


### search for job title and cost column
def stripDollarToInt(string):
  return float(string.split('$')[1])

jobTitleCol = 0
costCol = 0
minRow = 0


for row in readWbSheet.iter_rows():
  for cell in row:
    if type(cell.value) is str and cell.value.lower().strip() == 'job title':
      minRow = cell.row + 1
      # print(get_column_letter(cell.column))
      jobTitleCol = cell.column

    if type(cell.value) is str and cell.value.lower().strip() == 'cost':
      # print(get_column_letter(cell.column))
      # costCol = cell.column
      costCol = get_column_letter(cell.column)

if jobTitleCol == 0 or costCol == 0:
  print('did not find nesscary columns')

totalSum = 0

for row in readWbSheet.iter_rows(min_row = minRow, min_col = jobTitleCol, max_col=jobTitleCol):
  for cell in row:
    if cell.value == None:
      continue
    cost = stripDollarToInt(readWbSheet[costCol + str(cell.row)].value)
    totalSum += cost
    if(outputMap.get(cell.value, 'None') == 'None'):
      # print(jobTitleMap.get(cell.value,'None') == None)
      outputMap[cell.value] = {
        'company': None if jobTitleMap.get(cell.value,'None') == 'None' else jobTitleMap[cell.value]['company'],
        'code': None if jobTitleMap.get(cell.value,'None') == 'None' else jobTitleMap[cell.value]['code'],
        'cost': cost
      }
    # print(cell.value , readWbSheet[costCol + str(cell.row)].value)

# outputMap['invoiceNumber'] = invoiceNumber
# outputMap['totalCost'] = round(totalSum,2)

# print(outputMap)

# readWbSheet['J1'] = '=SUM(G9:G27)'
# readWb.save('test.xlsx')

######## create outputfile xlsx file #############
outputWb = openpyxl.Workbook()
outputWs = outputWb.active
outputWs.title = 'Sheet1'

rowCounter = 1

yellowFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

titleArray = ['Company', 'Job title', 'Cost', 'Code']

for idx, title in enumerate(titleArray):
  outputWs.cell(row=rowCounter, column=idx+1, value=title).border = thin_border

rowCounter += 1



for jobTitle, info in outputMap.items():
  # print(jobTitle,info)
  if info['company'] == None:
    outputWs.cell(row=rowCounter, column=1).fill = yellowFill
  else:
    outputWs.cell(row=rowCounter, column=1, value=info['company'])
  outputWs.cell(row=rowCounter, column=1).border = thin_border

  # if info['company'] == None:
  #   outputWs.cell(row=rowCounter, column=2).fill = yellowFill
  # else:
  outputWs.cell(row=rowCounter, column=2, value=jobTitle)
  outputWs.cell(row=rowCounter, column=2).border = thin_border

  if info['cost'] == None:
    outputWs.cell(row=rowCounter, column=3).fill = yellowFill
  else:
    outputWs.cell(row=rowCounter, column=3, value=info['cost'])
  outputWs.cell(row=rowCounter, column=3).border = thin_border

  if info['code'] == None:
    outputWs.cell(row=rowCounter, column=4).fill = yellowFill
  else:
    outputWs.cell(row=rowCounter, column=4, value=info['code'])
  outputWs.cell(row=rowCounter, column=4).border = thin_border

  rowCounter += 1

rowCounter += 1
outputWs.cell(row=rowCounter, column=1, value='Indeed invoice #' + str(invoiceNumber))
outputWs.cell(row=rowCounter, column=2, value='Total')
outputWs.cell(row=rowCounter, column=3, value=totalSum)

rowCounter += 2

outputWb.save(outputFileName)

####### tkinter functions ###########


def open():
  global fileNames
  root.fileNames = filedialog.askopenfilenames(initialdir="C:\\Users\\Terrence\\Projects\\python\\indeedInvoiceGenerator", title="Select Files", filetypes=(('xlsx files','*.xlsx'),("all files","*.*")))
  my_label = Label(root, text=root.fileNames).pack()

  print(type(root.fileNames))
  # InputFileList = root.tk.splitlist(root.fileNames)
  # print('Files = ', type(InputFileList))

chooseFilesButton = Button(root, text="Choose Excel Files", command=open).pack()

# root.mainloop()









