import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import csv
import os

class InvoiceGen:
  def __init__(self, fileNames, fileJobCodes, outputFileName):
    self.fileNames = fileNames
    self.fileJobCodes = fileJobCodes
    self.wbs = []
    self.jobTitleMap = {}

    self.resultMaps = []

    self.outputFileName = outputFileName
    self.currentDir = os.getcwd()

  def run(self):
    self.createJobCodeMap()
    self.createWbsFromCsvs()
    self.createResultMapFromCsvs()
    self.createResultExcelFromResultMaps()


  def getFileNames(self):
    print(self.fileNames)

  def createExcelFromCsv(self, fileName):
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(fileName) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)
    return wb
  
  def createWbsFromCsvs(self):
    for fn in self.fileNames:
      wb = self.createExcelFromCsv(fn)
      self.wbs.append(wb)
  
  def createJobCodeMap(self):
    jobCodeWb = openpyxl.load_workbook(filename = self.fileJobCodes)
    jobCodeSheet = jobCodeWb['Sheet1']

    for cell in jobCodeSheet['B']:
      if cell.row == 1 or cell.value.strip() == '':
        continue
      if(self.jobTitleMap.get(cell.value, 'None') == 'None'):
        self.jobTitleMap[cell.value] = {
          'company': jobCodeSheet['A' + str(cell.row)].value,
          'code': jobCodeSheet['C'+ str(cell.row)].value
        }
  
  def createResultMapFromCsvs(self):
    for wb in self.wbs:
      readWbSheet = wb['Sheet']
      invoiceNumber = readWbSheet['A1'].value.split('#')[1]

      jobTitleCol = 0
      costCol = 0
      minRow = 0
      totalSum = 0
      data = []
      outputMap = {}

      for row in readWbSheet.iter_rows():
        for cell in row:
          if type(cell.value) is str and cell.value.lower().strip() == 'job title':
            minRow = cell.row + 1
            jobTitleCol = cell.column

          if type(cell.value) is str and cell.value.lower().strip() == 'cost':
            costCol = get_column_letter(cell.column)

      if jobTitleCol == 0 or costCol == 0:
        print('did not find nesscary columns')
      
      for row in readWbSheet.iter_rows(min_row = minRow, min_col = jobTitleCol, max_col=jobTitleCol):
        for cell in row:
          if cell.value == None or cell.value.strip() == '':
            continue
          cost = self.stripDollarToInt(readWbSheet[costCol + str(cell.row)].value)
          totalSum += cost
          obj = {}

          obj['jobTitle'] = cell.value
          obj['company'] = None if self.jobTitleMap.get(cell.value,'None') == 'None' else self.jobTitleMap[cell.value]['company']
          obj['code'] = None if self.jobTitleMap.get(cell.value,'None') == 'None' else self.jobTitleMap[cell.value]['code']
          obj['cost'] = cost

          data.append(obj)
      
      outputMap['data'] = data
      outputMap['invoiceNumber'] = invoiceNumber
      outputMap['totalSum'] = totalSum

      self.resultMaps.append(outputMap)

  def createResultExcelFromResultMaps(self):
    outputWb = openpyxl.Workbook()
    outputWs = outputWb.active
    outputWs.title = 'Sheet1'

    rowCounter = 1
    columnWidths = [5] * 5

    yellowFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    for resultMap in self.resultMaps:
      titleArray = ['Company', 'Job title', 'Cost', 'Code']

      for idx, title in enumerate(titleArray):
        outputWs.cell(row=rowCounter, column=idx+1, value=title).border = thin_border

      rowCounter += 1

      for info in resultMap['data']:
        if info['company'] == None:
          outputWs.cell(row=rowCounter, column=1).fill = yellowFill
        else:
          outputWs.cell(row=rowCounter, column=1, value=info['company'])
          columnWidths[1] = max(columnWidths[1], len(info['company']))
        outputWs.cell(row=rowCounter, column=1).border = thin_border
        

        outputWs.cell(row=rowCounter, column=2, value=info['jobTitle'])
        outputWs.cell(row=rowCounter, column=2).border = thin_border
        columnWidths[2] = max(columnWidths[2], len(info['jobTitle']))

        if info['cost'] == None:
          outputWs.cell(row=rowCounter, column=3).fill = yellowFill
        else:
          outputWs.cell(row=rowCounter, column=3, value=info['cost']).number_format = '$#,##0.00'
          columnWidths[3] = max(columnWidths[3], len(str(info['cost'])))
        outputWs.cell(row=rowCounter, column=3).border = thin_border

        if info['code'] == None:
          outputWs.cell(row=rowCounter, column=4).fill = yellowFill
        else:
          outputWs.cell(row=rowCounter, column=4, value=info['code'])
          columnWidths[4] = max(columnWidths[4], len(info['code']))
        outputWs.cell(row=rowCounter, column=4).border = thin_border

        rowCounter += 1

      rowCounter += 1
      outputWs.cell(row=rowCounter, column=1, value='Indeed invoice #' + str(resultMap['invoiceNumber'])).font = Font(bold=True)
      columnWidths[1] = max(columnWidths[1], len('Indeed invoice #' + str(resultMap['invoiceNumber'])))
      outputWs.cell(row=rowCounter, column=2, value='Total').font = Font(bold=True)

      outputWs.cell(row=rowCounter, column=3, value=resultMap['totalSum']).font = Font(bold=True)
      outputWs.cell(row=rowCounter, column=3).number_format = '$#,##0.00'
      columnWidths[3] = max(columnWidths[3], len(str(resultMap['totalSum']+1)))
      rowCounter += 4

    for idx, columnWidth in enumerate(columnWidths):
      if idx == 0:
        continue
      outputWs.column_dimensions[get_column_letter(idx)].width = columnWidth + 2

    counter = 0
    while self.fileExists(self.currentDir + '/Invoices/' + self.outputFileName + ('({})'.format(counter) if counter > 0 else '') + '.xls'):
      counter += 1

    outputWb.save(self.currentDir + '/Invoices/' + self.outputFileName + ('({})'.format(counter) if counter > 0 else '') + '.xls')
  
  def stripDollarToInt(self,string):
    return float(string.split('$')[1])
  
  def fileExists(self,fileName):
    return os.path.exists(fileName)
    

